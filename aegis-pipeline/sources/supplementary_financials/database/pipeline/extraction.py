"""Extract financial supplement XLSX workbooks into inspectable artifacts.

This stage runs after the manifest stage. It reads ``files_to_process.json`` and
the per-file artifact folders produced by ``manifest.py``, opens each pending
workbook, and writes a workbook artifact containing one JSON file per visible
sheet. It intentionally stops after extraction; chunking, enrichment, and final
record creation are separate future stages.
"""

from __future__ import annotations

import json
import re
import time
import warnings
from collections.abc import Callable, Mapping
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl.utils.cell import range_boundaries, range_to_tuple

from ..utils.config_setup import get_input_source_config, load_config
from ..utils.logging_setup import get_stage_logger
from .manifest import (
    ARTIFACTS_DIR_NAME,
    FILES_TO_PROCESS_FILE_NAME,
    PROGRESS_DIR,
    ManifestRecord,
)

EXTRACTION_DIR_NAME = "extraction"
WORKBOOK_ARTIFACT_FILE_NAME = "workbook.json"
EXTRACTION_MANIFEST_FILE_NAME = "extraction_manifest.json"
DEFAULT_SHEET_WORKERS = 4
VISUAL_SERIES_POINT_LIMIT = 12
SYSTEM_SHEET_TITLES = frozenset({"variables"})
MANIFEST_FIELDS = (
    "file_id",
    "data_source",
    "fiscal_year",
    "quarter",
    "bank",
    "file_path",
    "file_name",
    "file_type",
    "file_size",
    "file_hash",
    "date_last_modified",
)


class ExtractionStageError(RuntimeError):
    """Raised when the XLSX extraction stage cannot continue safely."""


@dataclass(frozen=True)
class XlsxCell:
    """One populated worksheet cell after formula-safe normalization."""

    row_number: int
    column_index: int
    column_letter: str
    address: str
    value: str
    raw_value: Any
    cached_value: Any
    number_format: str
    data_type: str
    is_formula: bool
    has_cached_value: bool
    style: dict[str, Any]

    def to_record(self) -> dict[str, Any]:
        """Return the cell as a JSON-serializable record."""
        return {
            "row_number": self.row_number,
            "column_index": self.column_index,
            "column_letter": self.column_letter,
            "address": self.address,
            "value": self.value,
            "raw_value": self.raw_value,
            "cached_value": self.cached_value,
            "number_format": self.number_format,
            "data_type": self.data_type,
            "is_formula": self.is_formula,
            "has_cached_value": self.has_cached_value,
            "style": self.style,
        }


@dataclass(frozen=True)
class XlsxCellGrid:
    """Sparse worksheet grid containing only populated cells."""

    columns: list[int]
    rows: dict[int, dict[int, XlsxCell]]

    def to_record(self) -> dict[str, Any]:
        """Return the grid as a JSON-serializable record."""
        return {
            "columns": [
                {"index": column, "letter": _column_letter(column)}
                for column in self.columns
            ],
            "rows": [
                {
                    "row_number": row_number,
                    "cells": [
                        cell.to_record() for _column, cell in sorted(row_cells.items())
                    ],
                }
                for row_number, row_cells in sorted(self.rows.items())
            ],
        }


@dataclass(frozen=True)
class XlsxChartSeries:
    """Recovered chart series metadata."""

    name: str
    points: list[str]
    name_reference: str = ""
    category_reference: str = ""
    value_reference: str = ""

    def to_record(self) -> dict[str, Any]:
        """Return the series as a JSON-serializable record."""
        return {
            "name": self.name,
            "points": self.points,
            "name_reference": self.name_reference,
            "category_reference": self.category_reference,
            "value_reference": self.value_reference,
        }


@dataclass(frozen=True)
class XlsxChart:
    """Recovered chart metadata from a worksheet."""

    chart_index: int
    chart_type: str
    title: str
    anchor_cell: str
    x_axis_title: str
    y_axis_title: str
    series: list[XlsxChartSeries]

    def markdown(self, sheet_name: str) -> str:
        """Render chart metadata in the same markdown shape as aegis-retriever."""
        header = f'> [Chart]: {self.chart_type} - "{self.title}"'
        context_parts = []
        if self.x_axis_title:
            context_parts.append(f"X-axis: {self.x_axis_title}")
        if self.y_axis_title:
            context_parts.append(f"Y-axis: {self.y_axis_title}")
        context_parts.append(f"Sheet: {sheet_name}, Cell: {self.anchor_cell}")

        lines = [header, f"> {', '.join(context_parts)}"]
        for series in self.series:
            if series.points:
                preview_points = series.points[:VISUAL_SERIES_POINT_LIMIT]
                suffix = (
                    f" (+{len(series.points) - len(preview_points)} more)"
                    if len(series.points) > len(preview_points)
                    else ""
                )
                lines.append(
                    f'> Series "{series.name}": {", ".join(preview_points)}{suffix}'
                )
            else:
                lines.append(f'> Series "{series.name}": no data points recovered')
        return "\n".join(lines)

    def to_record(self) -> dict[str, Any]:
        """Return the chart as a JSON-serializable record."""
        return {
            "chart_index": self.chart_index,
            "chart_type": self.chart_type,
            "title": self.title,
            "anchor_cell": self.anchor_cell,
            "x_axis_title": self.x_axis_title,
            "y_axis_title": self.y_axis_title,
            "series": [series.to_record() for series in self.series],
        }


@dataclass(frozen=True)
class ExtractedXlsxSheet:
    """One extracted workbook sheet."""

    sheet_number: int
    title: str
    is_chartsheet: bool
    cell_grid: XlsxCellGrid | None
    charts: list[XlsxChart]
    images: list[dict[str, Any]]
    metadata: dict[str, Any]
    merged_ranges: list[dict[str, Any]]
    formula_stats: dict[str, int]
    markdown: str

    def to_record(self) -> dict[str, Any]:
        """Return the sheet content and metadata as a JSON record."""
        return {
            "sheet_number": self.sheet_number,
            "title": self.title,
            "is_chartsheet": self.is_chartsheet,
            "cell_grid": (
                self.cell_grid.to_record() if self.cell_grid is not None else None
            ),
            "charts": [chart.to_record() for chart in self.charts],
            "images": list(self.images),
            "metadata": self.metadata,
            "merged_ranges": self.merged_ranges,
            "formula_stats": self.formula_stats,
            "markdown": self.markdown,
        }


@dataclass(frozen=True)
class SkippedXlsxSheet:
    """Workbook sheet skipped before content extraction."""

    sheet_number: int
    title: str
    sheet_state: str
    reason: str

    def to_record(self) -> dict[str, Any]:
        """Return skipped-sheet metadata as a JSON record."""
        return {
            "sheet_number": self.sheet_number,
            "title": self.title,
            "sheet_state": self.sheet_state,
            "reason": self.reason,
        }


@dataclass(frozen=True)
class ExtractedXlsxWorkbook:
    """Extracted workbook content and sheet-selection metadata."""

    sheets: list[ExtractedXlsxSheet]
    source_sheet_count: int
    selected_sheet_numbers: list[int]
    skipped_sheets: list[SkippedXlsxSheet]


@dataclass(frozen=True)
class ExtractionStageResult:
    """Summary returned after writing workbook extraction artifacts."""

    processed_file_count: int
    workbook_artifact_paths: tuple[Path, ...]


@dataclass(frozen=True)
class XlsxFormulaStats:
    """Formula extraction counts for one worksheet."""

    formula_cell_count: int = 0
    formula_cached_value_count: int = 0
    formula_cache_missing_count: int = 0

    def to_record(self) -> dict[str, int]:
        """Return formula counts as a JSON-serializable record."""
        return {
            "formula_cell_count": self.formula_cell_count,
            "formula_cached_value_count": self.formula_cached_value_count,
            "formula_cache_missing_count": self.formula_cache_missing_count,
        }


@dataclass(frozen=True)
class XlsxCellExtraction:
    """Cells plus formula accounting collected from one worksheet."""

    cell_grid: XlsxCellGrid | None
    formula_stats: XlsxFormulaStats


@dataclass(frozen=True)
class _PreReadSheet:
    """Openpyxl-derived sheet data made safe for worker threads."""

    sheet_number: int
    title: str
    is_chartsheet: bool
    cell_grid: XlsxCellGrid | None
    charts: list[XlsxChart]
    metadata: dict[str, Any]
    merged_ranges: list[dict[str, Any]]
    formula_stats: dict[str, int]


def run_extraction_stage(
    input_base_path: Path | None = None,
    progress_dir: Path = PROGRESS_DIR,
    max_sheets: int | None = None,
    sheet_workers: int = DEFAULT_SHEET_WORKERS,
) -> ExtractionStageResult:
    """Extract all pending XLSX workbooks selected by the manifest stage.

    Args:
        input_base_path: Optional local financial-supp input folder override.
        progress_dir: Folder containing manifest progress files and artifacts.
        max_sheets: Optional visible-sheet limit for deterministic smoke runs.
        sheet_workers: Maximum worker threads used after sheets are pre-read.

    Returns:
        ExtractionStageResult with the workbook artifact paths written.

    Raises:
        ExtractionStageError: If progress files are invalid, source workbooks
            are missing, openpyxl is unavailable, or a workbook cannot be read.
        NotImplementedError: If configured input source is not local.

    External side effects:
        Writes artifacts under ``progress_dir/artifacts/<file_id>/extraction``.
    """
    logger = get_stage_logger(__name__, "EXTRACTION")
    input_base = _resolve_input_base_path(input_base_path)
    records = _load_process_records(progress_dir)
    processed_at = _utc_now()
    workbook_paths: list[Path] = []

    for index, record in enumerate(records, start=1):
        if record.file_type.lower() != "xlsx":
            raise ExtractionStageError(
                f"Extraction received non-xlsx file: {record.file_id}"
            )
        source_path = input_base / record.file_path
        if not source_path.is_file():
            raise ExtractionStageError(f"Source workbook is missing: {source_path}")

        logger.info(
            "Extracting workbook %d/%d: %s",
            index,
            len(records),
            record.file_name,
        )
        file_started_at = _utc_now()
        file_start = time.perf_counter()
        workbook = extract_xlsx_workbook(
            source_path,
            sheet_workers=sheet_workers,
            max_sheets=max_sheets,
        )
        duration_seconds = time.perf_counter() - file_start
        file_completed_at = _utc_now()
        artifact_path = write_workbook_artifact(
            record,
            source_path,
            workbook,
            _artifact_root(progress_dir, record.file_id) / EXTRACTION_DIR_NAME,
            processed_at,
            file_started_at=file_started_at,
            file_completed_at=file_completed_at,
            duration_seconds=duration_seconds,
        )
        workbook_paths.append(artifact_path)

    logger.info("Extraction complete: files=%d", len(records))
    return ExtractionStageResult(
        processed_file_count=len(records),
        workbook_artifact_paths=tuple(workbook_paths),
    )


def extract_xlsx_workbook(
    file_path: Path,
    sheet_workers: int = DEFAULT_SHEET_WORKERS,
    max_sheets: int | None = None,
) -> ExtractedXlsxWorkbook:
    """Extract one workbook into structured sheet records and markdown.

    The workbook is opened twice: once with formulas intact and once with
    cached displayed values. Formula text is never emitted. Cached values are
    used when present, and formula cells without cached values are skipped.
    """
    workbook, cached_workbook = _open_workbooks(file_path)
    try:
        source_sheet_count = len(workbook.sheetnames)
        sheet_items, skipped_sheets = _select_sheet_items(workbook, max_sheets)
        preread = _preread_selected_sheets(workbook, cached_workbook, sheet_items)
    finally:
        workbook.close()
        cached_workbook.close()

    sheets = _process_preread_sheets(
        preread,
        sheet_workers=sheet_workers,
    )
    return ExtractedXlsxWorkbook(
        sheets=sheets,
        source_sheet_count=source_sheet_count,
        selected_sheet_numbers=[sheet_number for sheet_number, _name in sheet_items],
        skipped_sheets=skipped_sheets,
    )


def write_workbook_artifact(
    record: ManifestRecord,
    source_path: Path,
    workbook: ExtractedXlsxWorkbook,
    extraction_root: Path,
    processed_at: str,
    file_started_at: str | None = None,
    file_completed_at: str | None = None,
    duration_seconds: float | None = None,
) -> Path:
    """Write the workbook artifact and one JSON file per extracted sheet."""
    extraction_root.mkdir(parents=True, exist_ok=True)
    source = _source_record(record, source_path)
    sheet_records = []
    for sheet in workbook.sheets:
        sheet_records.append(
            _write_sheet_json(source, sheet, extraction_root / "sheets", processed_at)
        )

    skipped_records = [sheet.to_record() for sheet in workbook.skipped_sheets]
    qa_findings = _xlsx_qa_findings(sheet_records)
    workbook_artifact = {
        "stage": "extraction",
        "processed_at": processed_at,
        "file_started_at": file_started_at or processed_at,
        "file_completed_at": file_completed_at or processed_at,
        "duration_seconds": round(float(duration_seconds or 0.0), 6),
        "source": source,
        "sheet_count": len(workbook.sheets),
        "source_sheet_count": workbook.source_sheet_count,
        "selected_sheet_numbers": workbook.selected_sheet_numbers,
        "extracted_sheet_count": len(workbook.sheets),
        "skipped_sheet_count": len(skipped_records),
        "skipped_sheets": skipped_records,
        "chart_count": sum(sheet["chart_count"] for sheet in sheet_records),
        "image_count": sum(sheet["image_count"] for sheet in sheet_records),
        "formula_stats": _aggregate_formula_stats(sheet_records),
        "qa_status": _qa_status(qa_findings),
        "qa_counts": _qa_counts(qa_findings),
        "qa_findings": qa_findings,
        "sheets": sheet_records,
    }
    workbook_path = extraction_root / WORKBOOK_ARTIFACT_FILE_NAME
    _write_json(workbook_path, workbook_artifact)
    _write_extraction_manifest(extraction_root, workbook_artifact)
    return workbook_path


def _write_sheet_json(
    source: dict[str, Any],
    sheet: ExtractedXlsxSheet,
    sheets_dir: Path,
    processed_at: str,
) -> dict[str, Any]:
    """Write one extracted sheet JSON artifact and return its index record."""
    sheet_label = f"sheet_{sheet.sheet_number:03d}"
    sheet_json_path = sheets_dir / f"{sheet_label}.json"
    payload = {
        "stage": "extraction_sheet",
        "processed_at": processed_at,
        "source": source,
        **sheet.to_record(),
    }
    _write_json(sheet_json_path, payload)
    return {
        "sheet_number": sheet.sheet_number,
        "sheet_title": sheet.title,
        "sheet_json_path": str(sheet_json_path),
        "row_count": len(sheet.cell_grid.rows) if sheet.cell_grid is not None else 0,
        "chart_count": len(sheet.charts),
        "image_count": len(sheet.images),
        "used_range": sheet.metadata.get("used_range", ""),
        "merged_range_count": len(sheet.merged_ranges),
        "formula_stats": sheet.formula_stats,
        "hidden_row_count": sheet.metadata.get("hidden_row_count", 0),
        "hidden_column_count": sheet.metadata.get("hidden_column_count", 0),
    }


def _write_extraction_manifest(
    extraction_root: Path,
    workbook_artifact: Mapping[str, Any],
) -> None:
    """Write a compact extraction-stage manifest next to workbook.json."""
    _write_json(
        extraction_root / EXTRACTION_MANIFEST_FILE_NAME,
        {
            "stage": workbook_artifact["stage"],
            "processed_at": workbook_artifact["processed_at"],
            "source": workbook_artifact["source"],
            "workbook_artifact_path": str(
                extraction_root / WORKBOOK_ARTIFACT_FILE_NAME
            ),
            "duration_seconds": workbook_artifact["duration_seconds"],
            "sheet_count": workbook_artifact["sheet_count"],
            "qa_status": workbook_artifact["qa_status"],
        },
    )


def _open_workbooks(file_path: Path) -> tuple[Any, Any]:
    """Open formula and cached-value workbooks with openpyxl."""
    load_workbook = _load_openpyxl()
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Print area cannot be set to Defined name:*",
            category=UserWarning,
            module="openpyxl.reader.workbook",
        )
        try:
            workbook = load_workbook(filename=file_path, data_only=False)
        except (OSError, ValueError) as exc:
            raise ExtractionStageError(
                f"Failed to open XLSX '{file_path.name}': {exc}"
            ) from exc
        success = False
        try:
            try:
                cached_workbook = load_workbook(filename=file_path, data_only=True)
            except (OSError, ValueError) as exc:
                raise ExtractionStageError(
                    f"Failed to open cached XLSX '{file_path.name}': {exc}"
                ) from exc
            success = True
            return workbook, cached_workbook
        finally:
            if not success:
                workbook.close()


def _load_openpyxl() -> Callable[..., Any]:
    """Return openpyxl.load_workbook or raise an actionable dependency error."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ExtractionStageError(
            "XLSX extraction requires openpyxl. Install it in the project venv "
            "before running real workbook extraction."
        ) from exc
    return load_workbook


def _select_sheet_items(
    workbook: Any,
    max_sheets: int | None,
) -> tuple[list[tuple[int, str]], list[SkippedXlsxSheet]]:
    """Return visible sheets selected for extraction and skipped metadata."""
    visible_sheets = []
    skipped_sheets = []
    for sheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        sheet_state = str(getattr(sheet, "sheet_state", "visible"))
        if sheet_state != "visible":
            skipped_sheets.append(
                SkippedXlsxSheet(
                    sheet_number=sheet_number,
                    title=sheet.title,
                    sheet_state=sheet_state,
                    reason="hidden_sheet",
                )
            )
            continue
        if max_sheets is not None and len(visible_sheets) >= max_sheets:
            skipped_sheets.append(
                SkippedXlsxSheet(
                    sheet_number=sheet_number,
                    title=sheet.title,
                    sheet_state=sheet_state,
                    reason="max_sheets_limit",
                )
            )
            continue
        visible_sheets.append((sheet_number, sheet_name))
    return visible_sheets, skipped_sheets


def _preread_selected_sheets(
    workbook: Any,
    cached_workbook: Any,
    sheet_items: list[tuple[int, str]],
) -> list[_PreReadSheet]:
    """Extract selected openpyxl sheet data into plain Python records."""
    preread = []
    for sheet_number, sheet_name in sheet_items:
        sheet = workbook[sheet_name]
        is_chartsheet = _is_chartsheet(sheet)
        cached_values = (
            {}
            if is_chartsheet or sheet_name not in cached_workbook.sheetnames
            else _build_cached_values(cached_workbook[sheet_name])
        )
        cell_extraction = (
            XlsxCellExtraction(None, XlsxFormulaStats())
            if is_chartsheet
            else _collect_sheet_cells(sheet, cached_values)
        )
        merged_ranges = (
            [] if is_chartsheet else _extract_merged_ranges(sheet, cached_values)
        )
        preread.append(
            _PreReadSheet(
                sheet_number=sheet_number,
                title=sheet.title,
                is_chartsheet=is_chartsheet,
                cell_grid=cell_extraction.cell_grid,
                charts=_extract_charts(sheet, workbook, cached_workbook),
                metadata=_extract_sheet_metadata(sheet, cell_extraction.cell_grid),
                merged_ranges=merged_ranges,
                formula_stats=cell_extraction.formula_stats.to_record(),
            )
        )
    return preread


def _process_preread_sheets(
    preread: list[_PreReadSheet],
    sheet_workers: int,
) -> list[ExtractedXlsxSheet]:
    """Build sheet markdown in parallel without touching openpyxl objects."""
    if not preread:
        return []
    worker_count = max(1, min(sheet_workers, len(preread)))
    if worker_count == 1:
        return [_process_one_sheet(sheet) for sheet in preread]

    results: list[ExtractedXlsxSheet | None] = [None] * len(preread)
    with ThreadPoolExecutor(max_workers=worker_count) as pool:
        futures = {
            pool.submit(_process_one_sheet, sheet): index
            for index, sheet in enumerate(preread)
        }
        for future in as_completed(futures):
            results[futures[future]] = future.result()
    return [result for result in results if result is not None]


def _process_one_sheet(sheet: _PreReadSheet) -> ExtractedXlsxSheet:
    """Build markdown for one pre-read sheet."""
    markdown = _assemble_sheet_markdown(
        sheet.title,
        sheet.cell_grid,
        sheet.charts,
    )
    return ExtractedXlsxSheet(
        sheet_number=sheet.sheet_number,
        title=sheet.title,
        is_chartsheet=sheet.is_chartsheet,
        cell_grid=sheet.cell_grid,
        charts=sheet.charts,
        images=[],
        metadata=sheet.metadata,
        merged_ranges=sheet.merged_ranges,
        formula_stats=sheet.formula_stats,
        markdown=markdown,
    )


def _build_cached_values(cached_sheet: Any) -> dict[tuple[int, int], Any]:
    """Build a lookup of cached displayed cell values."""
    values = {}
    for row in cached_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[(cell.row, cell.column)] = cell.value
    return values


def _extract_sheet_metadata(
    sheet: Any,
    cell_grid: XlsxCellGrid | None,
) -> dict[str, Any]:
    """Extract sheet-level layout metadata that should not inflate markdown."""
    hidden_rows = [
        int(row_number)
        for row_number, dimension in getattr(sheet, "row_dimensions", {}).items()
        if getattr(dimension, "hidden", False)
    ]
    hidden_columns = [
        str(column_letter)
        for column_letter, dimension in getattr(sheet, "column_dimensions", {}).items()
        if getattr(dimension, "hidden", False)
    ]
    tab_color = getattr(getattr(sheet, "sheet_properties", None), "tabColor", None)
    used_range = ""
    if cell_grid is not None:
        used_range = _used_range_from_grid(cell_grid)
    elif hasattr(sheet, "calculate_dimension"):
        try:
            used_range = str(sheet.calculate_dimension())
        except ValueError:
            used_range = ""
    return {
        "sheet_state": str(getattr(sheet, "sheet_state", "visible")),
        "used_range": used_range,
        "max_row": int(getattr(sheet, "max_row", 0) or 0),
        "max_column": int(getattr(sheet, "max_column", 0) or 0),
        "freeze_panes": str(getattr(sheet, "freeze_panes", "") or ""),
        "print_area": _json_safe_value(getattr(sheet, "print_area", "")),
        "hidden_rows": hidden_rows,
        "hidden_row_count": len(hidden_rows),
        "hidden_columns": hidden_columns,
        "hidden_column_count": len(hidden_columns),
        "tab_color": str(getattr(tab_color, "rgb", "") or ""),
    }


def _extract_merged_ranges(
    sheet: Any,
    cached_values: dict[tuple[int, int], Any],
) -> list[dict[str, Any]]:
    """Extract merged-cell ranges and their top-left displayed values."""
    merged = []
    merged_cells = getattr(sheet, "merged_cells", None)
    ranges = getattr(merged_cells, "ranges", []) if merged_cells is not None else []
    for merged_range in ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = sheet.cell(row=min_row, column=min_col)
        is_formula = isinstance(top_left.value, str) and top_left.value.startswith("=")
        display_value = cached_values.get(
            (top_left.row, top_left.column),
            "" if is_formula else top_left.value,
        )
        merged.append(
            {
                "range": str(merged_range),
                "top_left": f"{_column_letter(min_col)}{min_row}",
                "min_row": min_row,
                "min_column": min_col,
                "max_row": max_row,
                "max_column": max_col,
                "value": _normalize_cell_value(
                    display_value,
                    top_left.number_format or "",
                ),
            }
        )
    return merged


def _collect_populated_cells(
    sheet: Any,
    cached_values: dict[tuple[int, int], Any],
) -> XlsxCellGrid | None:
    """Collect non-empty worksheet cells into a sparse grid."""
    return _collect_sheet_cells(sheet, cached_values).cell_grid


def _collect_sheet_cells(
    sheet: Any,
    cached_values: dict[tuple[int, int], Any],
) -> XlsxCellExtraction:
    """Collect worksheet cells and formula-cache accounting."""
    rows: dict[int, dict[int, XlsxCell]] = {}
    populated_columns: set[int] = set()
    formula_cell_count = 0
    formula_cached_value_count = 0
    formula_cache_missing_count = 0
    for row in sheet.iter_rows():
        row_cells = {}
        for cell in row:
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            cached_value = cached_values.get((cell.row, cell.column))
            has_cached_value = (cell.row, cell.column) in cached_values
            if is_formula:
                formula_cell_count += 1
                if has_cached_value and cached_value is not None:
                    formula_cached_value_count += 1
                else:
                    formula_cache_missing_count += 1
            display_value = cached_values.get(
                (cell.row, cell.column),
                "" if is_formula else cell.value,
            )
            normalized_value = _normalize_cell_value(
                display_value,
                cell.number_format or "",
            )
            if not normalized_value:
                continue
            row_cells[cell.column] = XlsxCell(
                row_number=cell.row,
                column_index=cell.column,
                column_letter=_column_letter(cell.column),
                address=f"{_column_letter(cell.column)}{cell.row}",
                value=normalized_value,
                raw_value=None if is_formula else _json_safe_value(cell.value),
                cached_value=(
                    _json_safe_value(cached_value)
                    if is_formula and has_cached_value
                    else None
                ),
                number_format=cell.number_format or "",
                data_type=str(getattr(cell, "data_type", "")),
                is_formula=is_formula,
                has_cached_value=bool(is_formula and has_cached_value),
                style=_extract_cell_style(cell),
            )
            populated_columns.add(cell.column)
        if row_cells:
            rows[row[0].row] = row_cells
    formula_stats = XlsxFormulaStats(
        formula_cell_count=formula_cell_count,
        formula_cached_value_count=formula_cached_value_count,
        formula_cache_missing_count=formula_cache_missing_count,
    )
    if not rows:
        return XlsxCellExtraction(cell_grid=None, formula_stats=formula_stats)
    return XlsxCellExtraction(
        cell_grid=XlsxCellGrid(columns=sorted(populated_columns), rows=rows),
        formula_stats=formula_stats,
    )


def _normalize_cell_value(value: Any, number_format: str = "") -> str:
    """Normalize a worksheet cell value into stable text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if number_format and "%" in number_format:
            return _format_float(value * 100) + "%"
        return _format_float(value)
    if isinstance(value, int):
        return str(value)
    return _normalize_text_value(str(value))


def _json_safe_value(value: Any) -> Any:
    """Return a JSON-safe scalar without exposing formula strings."""
    if value is None:
        return None
    if isinstance(value, bool | int | str):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, tuple | list):
        return [_json_safe_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_safe_value(item) for key, item in value.items()}
    return str(value)


def _extract_cell_style(cell: Any) -> dict[str, Any]:
    """Extract compact semantic style hints for audit/debug JSON."""
    style: dict[str, Any] = {}
    font = getattr(cell, "font", None)
    if getattr(font, "bold", False):
        style["bold"] = True
    if getattr(font, "italic", False):
        style["italic"] = True

    alignment = getattr(cell, "alignment", None)
    indent = getattr(alignment, "indent", 0) or 0
    if indent:
        style["indent"] = float(indent)
    horizontal = getattr(alignment, "horizontal", None)
    if horizontal:
        style["horizontal_alignment"] = str(horizontal)

    fill = getattr(cell, "fill", None)
    fill_type = getattr(fill, "fill_type", None)
    if fill_type:
        style["fill_type"] = str(fill_type)
        fill_color = _color_value(getattr(fill, "fgColor", None))
        if fill_color:
            style["fill_color"] = fill_color

    border = getattr(cell, "border", None)
    border_edges = [
        edge_name
        for edge_name in ("left", "right", "top", "bottom")
        if getattr(getattr(border, edge_name, None), "style", None)
    ]
    if border_edges:
        style["border_edges"] = border_edges

    return style


def _color_value(color: Any) -> str:
    """Return a compact color value from an openpyxl color object."""
    if color is None:
        return ""
    color_type = str(getattr(color, "type", "") or "")
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        if isinstance(rgb, str) and _looks_like_color_token(rgb):
            return rgb
        return ""
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        return f"indexed:{indexed}" if type(indexed) is int else ""
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        return f"theme:{theme}" if type(theme) is int else ""
    if color_type == "auto":
        return "auto"

    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and _looks_like_color_token(rgb):
        return rgb
    indexed = getattr(color, "indexed", None)
    if type(indexed) is int:
        return f"indexed:{indexed}"
    theme = getattr(color, "theme", None)
    if type(theme) is int:
        return f"theme:{theme}"
    return ""


def _looks_like_color_token(value: str) -> bool:
    """Return whether an openpyxl color token is useful to serialize."""
    return bool(re.fullmatch(r"[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8}", value))


def _used_range_from_grid(cell_grid: XlsxCellGrid) -> str:
    """Build a populated-cell used range from the sparse grid."""
    min_row = min(cell_grid.rows)
    max_row = max(cell_grid.rows)
    min_column = min(cell_grid.columns)
    max_column = max(cell_grid.columns)
    return (
        f"{_column_letter(min_column)}{min_row}:"
        f"{_column_letter(max_column)}{max_row}"
    )


def _format_float(value: float) -> str:
    """Return a compact decimal string rounded to at most two places."""
    try:
        decimal_value = Decimal(str(value)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    except (InvalidOperation, ValueError):
        return str(value)
    if decimal_value == 0:
        decimal_value = Decimal("0")
    return format(decimal_value, "f").rstrip("0").rstrip(".")


def _normalize_text_value(value: str) -> str:
    """Normalize Excel text escapes into retrievable plain text."""
    return (
        re.sub(r"_x000[Dd]_", "\n", value)
        .replace("\r\n", "\n")
        .replace("\r", "\n")
    )


def _assemble_sheet_markdown(
    sheet_name: str,
    cell_grid: XlsxCellGrid | None,
    charts: list[XlsxChart],
) -> str:
    """Combine cell markdown and chart descriptions for one sheet."""
    visual_blocks = [chart.markdown(sheet_name) for chart in charts]
    if cell_grid is None and not visual_blocks:
        return f"# Sheet: {sheet_name}\n\nThis sheet contains no data."
    if cell_grid is None:
        return "\n\n".join([f"# Sheet: {sheet_name}", *visual_blocks])
    sheet_markdown = _serialize_sheet(sheet_name, cell_grid)
    if visual_blocks:
        return "\n\n".join([sheet_markdown, *visual_blocks])
    return sheet_markdown


def _serialize_sheet(sheet_name: str, cell_grid: XlsxCellGrid) -> str:
    """Serialize sparse cells into row-numbered markdown."""
    header_cells = [_column_letter(column) for column in cell_grid.columns]
    lines = [
        f"# Sheet: {sheet_name}",
        "",
        f"| Row | {' | '.join(header_cells)} |",
        f"| {' | '.join(['---'] * (len(header_cells) + 1))} |",
    ]
    for row_number, row_cells in sorted(cell_grid.rows.items()):
        rendered_row = [
            _escape_table_text(row_cells[column].value) if column in row_cells else ""
            for column in cell_grid.columns
        ]
        lines.append(f"| {row_number} | {' | '.join(rendered_row)} |")
    return "\n".join(lines)


def _extract_charts(sheet: Any, workbook: Any, cached_workbook: Any) -> list[XlsxChart]:
    """Extract chart metadata from one worksheet."""
    charts = []
    for index, chart in enumerate(getattr(sheet, "_charts", []), start=1):
        x_axis = getattr(chart, "x_axis", None)
        y_axis = getattr(chart, "y_axis", None)
        series = [
            XlsxChartSeries(
                name=_extract_series_name(
                    item,
                    workbook,
                    cached_workbook,
                    series_index,
                ),
                points=_extract_series_points(item, workbook, cached_workbook),
                name_reference=_extract_series_name_reference(item),
                category_reference=_extract_series_category_reference(item),
                value_reference=_extract_series_value_reference(item),
            )
            for series_index, item in enumerate(getattr(chart, "ser", []), start=1)
        ]
        charts.append(
            XlsxChart(
                chart_index=index,
                chart_type=type(chart).__name__,
                title=_extract_title_text(getattr(chart, "title", None))
                or "Untitled Chart",
                anchor_cell=_extract_anchor_cell(getattr(chart, "anchor", None)),
                x_axis_title=_extract_title_text(getattr(x_axis, "title", None)),
                y_axis_title=_extract_title_text(getattr(y_axis, "title", None)),
                series=series,
            )
        )
    return charts


def _extract_title_text(title: Any) -> str:
    """Extract plain text from an openpyxl title-like object."""
    if title is None:
        return ""
    if isinstance(title, str):
        return title.strip()
    text_runs = []
    text = getattr(title, "tx", None)
    rich_text = getattr(text, "rich", None)
    paragraphs = getattr(rich_text, "p", [])
    for paragraph in paragraphs:
        for run in getattr(paragraph, "r", []) or []:
            run_text = getattr(run, "t", "")
            if run_text:
                text_runs.append(str(run_text))
        for fld_item in getattr(paragraph, "fld", []) or []:
            field_text = getattr(fld_item, "t", "")
            if field_text:
                text_runs.append(str(field_text))
    return " ".join(part.strip() for part in text_runs if part).strip()


def _extract_anchor_cell(anchor: Any) -> str:
    """Extract a drawing anchor cell."""
    marker = getattr(anchor, "_from", None)
    if marker is None and hasattr(anchor, "from_"):
        marker = getattr(anchor, "from_", None)
    if marker is not None:
        row = getattr(marker, "row", None)
        column = getattr(marker, "col", None)
        if isinstance(row, int) and isinstance(column, int):
            return f"{_column_letter(column + 1)}{row + 1}"
    if isinstance(anchor, str):
        parsed = _parse_cell_range(anchor)
        if parsed is not None:
            _sheet_name, min_col, min_row, _max_col, _max_row = parsed
            return f"{_column_letter(min_col)}{min_row}"
    return "unknown"


def _extract_series_name(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
    series_index: int,
) -> str:
    """Extract a human-readable chart series name."""
    text_source = getattr(series, "tx", None)
    literal_value = getattr(text_source, "v", None)
    if isinstance(literal_value, str) and literal_value.strip():
        return literal_value.strip()
    string_reference = getattr(text_source, "strRef", None)
    formula = getattr(string_reference, "f", "")
    for value in _load_reference_values(workbook, cached_workbook, formula):
        if value:
            return value
    return f"Series {series_index}"


def _extract_series_name_reference(series: Any) -> str:
    """Return the source range for a chart series name, when present."""
    text_source = getattr(series, "tx", None)
    string_reference = getattr(text_source, "strRef", None)
    return str(getattr(string_reference, "f", "") or "")


def _extract_series_category_reference(series: Any) -> str:
    """Return the source range for chart categories, when present."""
    category_source = getattr(series, "cat", None)
    category_reference = getattr(category_source, "strRef", None)
    if category_reference is None:
        category_reference = getattr(category_source, "numRef", None)
    return str(getattr(category_reference, "f", "") or "")


def _extract_series_value_reference(series: Any) -> str:
    """Return the source range for chart values, when present."""
    value_source = getattr(series, "val", None)
    value_reference = getattr(value_source, "numRef", None)
    return str(getattr(value_reference, "f", "") or "")


def _extract_series_points(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
) -> list[str]:
    """Extract paired category/value points for one chart series."""
    categories = _load_reference_values(
        workbook,
        cached_workbook,
        _extract_series_category_reference(series),
    )
    values = _load_reference_values(
        workbook,
        cached_workbook,
        _extract_series_value_reference(series),
    )
    points = []
    for index, (category, value) in enumerate(
        zip_longest(categories, values, fillvalue=""),
        start=1,
    ):
        if not category and not value:
            continue
        points.append(f"{category or f'Point {index}'}: {value}")
    return points


def _load_reference_values(
    workbook: Any,
    cached_workbook: Any,
    reference: str,
) -> list[str]:
    """Resolve an Excel reference into normalized cached values."""
    parsed = _parse_cell_range(reference)
    if parsed is None:
        return []
    sheet_name, min_col, min_row, max_col, max_row = parsed
    if sheet_name not in workbook.sheetnames:
        return []
    formula_sheet = workbook[sheet_name]
    cached_sheet = (
        cached_workbook[sheet_name]
        if sheet_name in cached_workbook.sheetnames
        else None
    )
    values = []
    for row_number in range(min_row, max_row + 1):
        for column_number in range(min_col, max_col + 1):
            formula_cell = formula_sheet.cell(row=row_number, column=column_number)
            raw_value = (
                cached_sheet.cell(row=row_number, column=column_number).value
                if cached_sheet is not None
                else formula_cell.value
            )
            if isinstance(raw_value, str) and raw_value.startswith("="):
                raw_value = ""
            values.append(
                _normalize_cell_value(
                    raw_value,
                    formula_cell.number_format or "",
                )
            )
    return values


def _xlsx_qa_findings(sheet_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return deterministic workbook QA findings from sheet records."""
    findings = []
    for record in sheet_records:
        formula_stats = record.get("formula_stats") or {}
        formula_cache_missing_count = int(
            formula_stats.get("formula_cache_missing_count", 0)
        )
        if formula_cache_missing_count:
            findings.append(
                {
                    "code": "xlsx_formula_cache_missing",
                    "severity": "medium",
                    "message": (
                        f"{formula_cache_missing_count} formula cell(s) omitted "
                        "because cached values were unavailable"
                    ),
                    "sheet_number": int(record["sheet_number"]),
                    "artifact_path": str(record.get("sheet_json_path", "")),
                    "formula_cache_missing_count": formula_cache_missing_count,
                }
            )

        sheet_title = str(record.get("sheet_title", "")).strip()
        if sheet_title.lower() not in SYSTEM_SHEET_TITLES:
            continue
        findings.append(
            {
                "code": "xlsx_system_sheet_emitted",
                "severity": "high",
                "message": f"System workbook sheet emitted as content: {sheet_title}",
                "sheet_number": int(record["sheet_number"]),
                "artifact_path": str(record.get("sheet_json_path", "")),
            }
        )
    return findings


def _aggregate_formula_stats(sheet_records: list[dict[str, Any]]) -> dict[str, int]:
    """Aggregate formula-cache accounting from sheet index records."""
    totals = {
        "formula_cell_count": 0,
        "formula_cached_value_count": 0,
        "formula_cache_missing_count": 0,
    }
    for record in sheet_records:
        formula_stats = record.get("formula_stats") or {}
        for key in totals:
            totals[key] += int(formula_stats.get(key, 0))
    return totals


def _qa_counts(findings: list[dict[str, Any]]) -> dict[str, int]:
    """Count QA findings by supported severity."""
    counts = {"high": 0, "low": 0, "medium": 0}
    for finding in findings:
        severity = str(finding.get("severity", "")).lower()
        if severity in counts:
            counts[severity] += 1
    return counts


def _qa_status(findings: list[dict[str, Any]]) -> str:
    """Return failed when high-severity findings are present."""
    return (
        "failed"
        if any(finding.get("severity") == "high" for finding in findings)
        else "passed"
    )


def _is_chartsheet(sheet: Any) -> bool:
    """Return whether an openpyxl sheet is a chartsheet."""
    return type(sheet).__name__ == "Chartsheet"


def _escape_table_text(value: str) -> str:
    """Escape markdown table delimiters in cell text."""
    return value.replace("|", "\\|").replace("\n", "<br>")


def _parse_cell_range(reference: str) -> tuple[str, int, int, int, int] | None:
    """Parse an Excel sheet range reference with openpyxl utilities."""
    if not reference:
        return None
    try:
        sheet_name, boundaries = range_to_tuple(reference)
    except ValueError:
        try:
            boundaries = range_boundaries(reference)
        except ValueError:
            return None
        sheet_name = ""
    min_col, min_row, max_col, max_row = boundaries
    return sheet_name, min_col, min_row, max_col, max_row


def _column_letter(column_index: int) -> str:
    """Return the Excel column letter for a one-based column index."""
    if column_index < 1:
        raise ValueError("column_index must be >= 1")
    letters = []
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def _load_process_records(progress_dir: Path) -> tuple[ManifestRecord, ...]:
    """Load manifest records selected for processing."""
    path = progress_dir / FILES_TO_PROCESS_FILE_NAME
    payload = _read_required_json(path)
    rows = payload.get("files_to_process")
    if not isinstance(rows, list):
        raise ExtractionStageError(f"{path} is missing files_to_process list")
    return tuple(_manifest_record_from_mapping(row, path) for row in rows)


def _manifest_record_from_mapping(row: Any, source_path: Path) -> ManifestRecord:
    """Convert one progress JSON row into a manifest record."""
    if not isinstance(row, Mapping):
        raise ExtractionStageError(f"Progress row is not an object in {source_path}")
    missing = [field for field in MANIFEST_FIELDS if field not in row]
    if missing:
        raise ExtractionStageError(
            f"Progress row missing fields in {source_path}: {', '.join(missing)}"
        )
    return ManifestRecord(
        file_id=str(row["file_id"]),
        data_source=str(row["data_source"]),
        fiscal_year=str(row["fiscal_year"]),
        quarter=str(row["quarter"]),
        bank=str(row["bank"]),
        file_path=str(row["file_path"]),
        file_name=str(row["file_name"]),
        file_type=str(row["file_type"]),
        file_size=int(row["file_size"]),
        file_hash=str(row["file_hash"]),
        date_last_modified=str(row["date_last_modified"]),
    )


def _source_record(record: ManifestRecord, source_path: Path) -> dict[str, Any]:
    """Build source metadata shared by extraction artifacts."""
    return {
        "file_id": record.file_id,
        "data_source": record.data_source,
        "fiscal_year": record.fiscal_year,
        "quarter": record.quarter,
        "bank": record.bank,
        "file_path": record.file_path,
        "file_name": record.file_name,
        "file_type": record.file_type,
        "file_size": record.file_size,
        "file_hash": record.file_hash,
        "date_last_modified": record.date_last_modified,
        "period": f"{record.fiscal_year}_{record.quarter}",
        "source_absolute_path": str(source_path),
    }


def _resolve_input_base_path(input_base_path: Path | None) -> Path:
    """Resolve the local financial-supp input folder."""
    if input_base_path is not None:
        path = input_base_path.expanduser().resolve()
        if not path.is_dir():
            raise ExtractionStageError(f"Input base path is not a directory: {path}")
        return path
    load_config()
    input_config = get_input_source_config()
    if input_config.source != "local":
        raise NotImplementedError(
            "Extraction currently supports local input paths only."
        )
    configured_path = input_config.base_path
    if not isinstance(configured_path, Path):
        configured_path = Path(configured_path)
    return configured_path.expanduser().resolve()


def _artifact_root(progress_dir: Path, file_id: str) -> Path:
    """Return the per-file artifact root created by the manifest stage."""
    return progress_dir / ARTIFACTS_DIR_NAME / file_id


def _read_required_json(path: Path) -> Any:
    """Read required UTF-8 JSON and raise a stage error on failure."""
    if not path.is_file():
        raise ExtractionStageError(f"Required JSON artifact is missing: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ExtractionStageError(f"Invalid JSON artifact: {path}") from exc


def _write_json(path: Path, payload: Any) -> None:
    """Write deterministic UTF-8 JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _utc_now() -> str:
    """Return the current UTC time as an ISO string."""
    return datetime.now(tz=UTC).isoformat()
