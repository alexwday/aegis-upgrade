"""Generate browser-preview bytes for persisted source documents."""

from __future__ import annotations

import io
import re
import textwrap
import warnings
import xml.etree.ElementTree as ET
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Any

PDF_MIME_TYPE = "application/pdf"
HTML_MIME_TYPE = "text/html; charset=utf-8"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XML_MIME_TYPES = {"application/xml", "text/xml"}
PREVIEW_RENDERER_VERSION = "source_document_preview_v3"

_SECTION_MD = "MANAGEMENT DISCUSSION SECTION"
_SECTION_QA = "Q&A"
_WHITESPACE_RE = re.compile(r"\s+")
_PAGE_MARGIN = 42
_TABLE_BORDER_COLOR = "#d8dee6"


@dataclass(frozen=True)
class SourceDocumentPreview:
    """Generated preview payload ready to persist in Postgres."""

    preview_mime_type: str
    preview_bytes: bytes
    preview_metadata: dict[str, Any]


@dataclass(frozen=True)
class _TranscriptParticipant:
    """One FactSet transcript participant."""

    participant_id: str
    name: str
    participant_type: str = ""
    title: str = ""
    affiliation: str = ""
    affiliation_entity: str = ""

    @property
    def display_name(self) -> str:
        parts = [self.name or "Unknown Speaker"]
        if self.title and self.title not in parts:
            parts.append(self.title)
        if self.affiliation and self.affiliation not in parts:
            parts.append(self.affiliation)
        return ", ".join(parts)


@dataclass(frozen=True)
class _TranscriptSpeakerBlock:
    """One consecutive speaker block from a transcript body."""

    speaker_block_id: int
    section_name: str
    raw_section_name: str
    participant_id: str
    speaker_type_hint: str
    participant: _TranscriptParticipant
    paragraphs: list[str]

    @property
    def speaker_name(self) -> str:
        return self.participant.name or "Unknown Speaker"


@dataclass(frozen=True)
class _TranscriptDocument:
    """Parsed transcript document."""

    title: str
    transcript_date: str
    participants: dict[str, _TranscriptParticipant]
    blocks: list[_TranscriptSpeakerBlock]
    companies: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _WorkbookSheet:
    """One visible workbook sheet selected for preview rendering."""

    page_number: int
    workbook_sheet_number: int
    name: str
    is_chartsheet: bool
    sheet_state: str


@dataclass(frozen=True)
class TranscriptPreviewPage:
    """One generated transcript PDF page and its retrieval text."""

    page_number: int
    title: str
    markdown: str
    section_name: str
    speaker_block_ids: list[int]
    speakers: list[dict[str, str]]


@dataclass(frozen=True)
class TranscriptPreviewDocument:
    """Generated transcript preview plus page-aligned extraction records."""

    preview: SourceDocumentPreview
    pages: list[TranscriptPreviewPage]
    title: str
    transcript_date: str
    companies: list[str]


def build_source_document_preview(
    *,
    original_bytes: bytes,
    filename: str,
    file_type: str,
    mime_type: str,
    source_type: str = "",
    file_hash: str = "",
) -> SourceDocumentPreview:
    """Return deterministic preview bytes plus metadata for one source document."""
    normalized_type = (file_type or Path(filename).suffix.lstrip(".")).lower()
    normalized_mime = (mime_type or "").lower()
    metadata = _base_metadata(
        filename=filename,
        file_type=normalized_type,
        mime_type=normalized_mime,
        source_type=source_type,
        file_hash=file_hash,
    )

    if normalized_type == "pdf" or normalized_mime == PDF_MIME_TYPE:
        return SourceDocumentPreview(
            preview_mime_type=PDF_MIME_TYPE,
            preview_bytes=original_bytes,
            preview_metadata={
                **metadata,
                "preview_kind": "pdf",
                "page_model": "original_pdf_pages",
            },
        )
    if normalized_type == "xlsx" or normalized_mime == XLSX_MIME_TYPE:
        return _build_xlsx_preview(
            original_bytes=original_bytes,
            filename=filename,
            metadata=metadata,
        )
    if normalized_type == "xml" or normalized_mime in XML_MIME_TYPES:
        return _build_xml_transcript_preview(
            original_bytes=original_bytes,
            filename=filename,
            metadata=metadata,
        )

    raise ValueError(
        f"Unsupported source document preview type: {file_type or mime_type or filename}"
    )


def is_preview_current(
    *,
    stored_file_hash: str,
    expected_file_hash: str,
    preview_mime_type: str | None,
    has_preview_bytes: bool,
    preview_metadata: Mapping[str, Any] | None,
    preview_error: str | None,
) -> bool:
    """Return whether a stored preview matches the current renderer contract."""
    if str(stored_file_hash) != str(expected_file_hash):
        return False
    if preview_error:
        return False
    if not preview_mime_type or not has_preview_bytes:
        return False
    return preview_metadata_is_current(preview_metadata)


def preview_metadata_is_current(
    preview_metadata: Mapping[str, Any] | None,
) -> bool:
    """Return whether metadata was produced by the current preview renderer."""
    return bool(
        preview_metadata
        and preview_metadata.get("renderer_version") == PREVIEW_RENDERER_VERSION
    )


def preview_error_metadata(
    *,
    filename: str,
    file_type: str,
    mime_type: str,
    source_type: str = "",
    file_hash: str = "",
    error: BaseException,
) -> dict[str, Any]:
    """Return metadata for a failed preview generation attempt."""
    return {
        **_base_metadata(
            filename=filename,
            file_type=file_type,
            mime_type=mime_type,
            source_type=source_type,
            file_hash=file_hash,
        ),
        "preview_kind": "error",
        "error_type": type(error).__name__,
    }


def _base_metadata(
    *,
    filename: str,
    file_type: str,
    mime_type: str,
    source_type: str,
    file_hash: str,
) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "renderer_version": PREVIEW_RENDERER_VERSION,
        "filename": filename,
        "source_file_type": file_type,
        "source_mime_type": mime_type,
    }
    if source_type:
        metadata["source_type"] = source_type
    if file_hash:
        metadata["source_file_hash"] = file_hash
    return metadata


def _build_xlsx_preview(
    *,
    original_bytes: bytes,
    filename: str,
    metadata: dict[str, Any],
) -> SourceDocumentPreview:
    workbook, cached_workbook = _load_preview_workbooks(original_bytes)
    try:
        workbook_sheet_count = len(workbook.sheetnames)
        visible_sheets = _visible_workbook_sheets(workbook)
        preview_bytes, sheet_records = _render_workbook_html(
            workbook=workbook,
            cached_workbook=cached_workbook,
            filename=filename,
            visible_sheets=visible_sheets,
        )
    finally:
        workbook.close()
        cached_workbook.close()

    return SourceDocumentPreview(
        preview_mime_type=HTML_MIME_TYPE,
        preview_bytes=preview_bytes,
        preview_metadata={
            **metadata,
            "preview_kind": "xlsx_html",
            "page_model": "visible_sheet_html_sections",
            "render_engine": "openpyxl_html",
            "sheets": sheet_records,
            "visible_sheet_count": len(visible_sheets),
            "workbook_sheet_count": workbook_sheet_count,
        },
    )


def _load_preview_workbooks(original_bytes: bytes) -> tuple[Any, Any]:
    """Load formula and cached-value workbooks for HTML preview rendering."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Print area cannot be set to Defined name:*",
            category=UserWarning,
            module="openpyxl.reader.workbook",
        )
        workbook = load_workbook(
            io.BytesIO(original_bytes),
            data_only=False,
            read_only=False,
        )
        success = False
        try:
            cached_workbook = load_workbook(
                io.BytesIO(original_bytes),
                data_only=True,
                read_only=False,
            )
            success = True
            return workbook, cached_workbook
        finally:
            if not success:
                workbook.close()


def _visible_workbook_sheets(workbook: Any) -> list[_WorkbookSheet]:
    """Return visible sheets in visible order, matching XLSX retrieval pages."""
    visible_sheets: list[_WorkbookSheet] = []
    for workbook_sheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        sheet_state = str(getattr(sheet, "sheet_state", "visible"))
        if sheet_state != "visible":
            continue
        page_number = len(visible_sheets) + 1
        visible_sheets.append(
            _WorkbookSheet(
                page_number=page_number,
                workbook_sheet_number=workbook_sheet_number,
                name=str(sheet.title),
                is_chartsheet=not hasattr(sheet, "iter_rows"),
                sheet_state=sheet_state,
            )
        )
    return visible_sheets


def _render_workbook_html(
    *,
    workbook: Any,
    cached_workbook: Any,
    filename: str,
    visible_sheets: list[_WorkbookSheet],
) -> tuple[bytes, list[dict[str, Any]]]:
    """Render a workbook as one HTML document with sheet anchors."""
    sheet_records: list[dict[str, Any]] = []
    sections = []
    if not visible_sheets:
        sections.append(
            '<section class="sheet is-empty" id="sheet-1">'
            "<h2>No visible sheets</h2>"
            "<p>This workbook has no visible sheets to preview.</p>"
            "</section>"
        )
    for preview_sheet in visible_sheets:
        formula_sheet = workbook[preview_sheet.name]
        cached_sheet = (
            cached_workbook[preview_sheet.name]
            if preview_sheet.name in cached_workbook.sheetnames
            else None
        )
        section_html, record = _render_sheet_section(
            preview_sheet=preview_sheet,
            formula_sheet=formula_sheet,
            cached_sheet=cached_sheet,
        )
        sections.append(section_html)
        sheet_records.append(record)

    title = escape(filename)
    html = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>{_workbook_html_styles()}</style>
</head>
<body>
  <header class="workbook-header">
    <div>
      <h1>{title}</h1>
      <p>{len(sheet_records)} visible sheet{"s" if len(sheet_records) != 1 else ""}</p>
    </div>
    {_sheet_nav_html(sheet_records)}
  </header>
  <main>
    {"".join(sections)}
  </main>
</body>
</html>
"""
    return html.encode("utf-8"), sheet_records


def _sheet_nav_html(sheet_records: list[dict[str, Any]]) -> str:
    """Return compact anchor navigation for visible workbook sheets."""
    if not sheet_records:
        return ""
    links = [
        f'<a href="#{escape(str(sheet["anchor"]))}">'
        f'{int(sheet["page_number"])}. {escape(str(sheet["name"]))}</a>'
        for sheet in sheet_records
    ]
    return f'<nav class="sheet-nav">{"".join(links)}</nav>'


def _render_sheet_section(
    *,
    preview_sheet: _WorkbookSheet,
    formula_sheet: Any,
    cached_sheet: Any | None,
) -> tuple[str, dict[str, Any]]:
    """Render one worksheet to HTML and metadata."""
    anchor = f"sheet-{preview_sheet.page_number}"
    record = {
        "name": preview_sheet.name,
        "page_number": preview_sheet.page_number,
        "visible_sheet_number": preview_sheet.page_number,
        "workbook_sheet_number": preview_sheet.workbook_sheet_number,
        "anchor": anchor,
        "sheet_state": preview_sheet.sheet_state,
        "is_chartsheet": preview_sheet.is_chartsheet,
        "preview_page": preview_sheet.page_number,
    }
    if preview_sheet.is_chartsheet:
        return (
            f'<section class="sheet is-empty" id="{anchor}">'
            f"<h2>Sheet {preview_sheet.page_number}: {escape(preview_sheet.name)}</h2>"
            "<p>Chartsheet previews are not rendered in the HTML workbook view.</p>"
            "</section>",
            {**record, "row_count": 0, "column_count": 0},
        )

    table_html, row_count, column_count = _render_sheet_table(
        formula_sheet,
        cached_sheet,
    )
    return (
        f'<section class="sheet" id="{anchor}">'
        f"<h2>Sheet {preview_sheet.page_number}: {escape(preview_sheet.name)}</h2>"
        f'<div class="sheet-scroll">{table_html}</div>'
        "</section>",
        {**record, "row_count": row_count, "column_count": column_count},
    )


def _render_sheet_table(sheet: Any, cached_sheet: Any | None) -> tuple[str, int, int]:
    """Return an Excel-like HTML table for one worksheet."""
    max_row = int(getattr(sheet, "max_row", 0) or 0)
    max_column = int(getattr(sheet, "max_column", 0) or 0)
    visible_rows = [
        row_number
        for row_number in range(1, max_row + 1)
        if not _row_is_hidden(sheet, row_number)
    ]
    visible_columns = [
        column_number
        for column_number in range(1, max_column + 1)
        if not _column_is_hidden(sheet, column_number)
    ]
    if not visible_rows or not visible_columns:
        return '<p class="empty-sheet">This sheet has no visible cells.</p>', 0, 0

    merge_starts, merge_covered = _merged_cell_maps(
        sheet,
        visible_rows=visible_rows,
        visible_columns=visible_columns,
    )
    colgroup = "".join(
        f'<col style="width:{_column_width_px(sheet, column_number)}px">'
        for column_number in visible_columns
    )
    header_cells = "".join(
        f"<th>{escape(_column_letter(column_number))}</th>"
        for column_number in visible_columns
    )
    row_html = [
        f"<tr><th class=\"corner\"></th>{header_cells}</tr>",
    ]
    for row_number in visible_rows:
        row_style = _style_attr(_row_style(sheet, row_number))
        row_attr = f" {row_style}" if row_style else ""
        cells = [f"<th{row_attr}>{row_number}</th>"]
        for column_number in visible_columns:
            if (row_number, column_number) in merge_covered:
                continue
            cell = sheet.cell(row=row_number, column=column_number)
            value = _display_cell_value(cell, cached_sheet)
            attrs = [
                f'data-cell="{escape(_column_letter(column_number))}{row_number}"',
                _style_attr(_cell_style(cell)),
            ]
            merge = merge_starts.get((row_number, column_number))
            if merge is not None:
                rowspan, colspan = merge
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
            attrs_text = " ".join(attr for attr in attrs if attr)
            rendered_value = _html_cell_value(value)
            cells.append(f"<td {attrs_text}>{rendered_value}</td>")
        row_html.append(f"<tr>{''.join(cells)}</tr>")
    table = (
        '<table class="sheet-grid">'
        f"<colgroup><col>{colgroup}</colgroup>"
        f"<tbody>{''.join(row_html)}</tbody>"
        "</table>"
    )
    return table, len(visible_rows), len(visible_columns)


def _merged_cell_maps(
    sheet: Any,
    *,
    visible_rows: list[int],
    visible_columns: list[int],
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    """Return merged-cell starts and covered visible coordinates."""
    visible_row_set = set(visible_rows)
    visible_column_set = set(visible_columns)
    starts: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    merged_cells = getattr(sheet, "merged_cells", None)
    ranges = getattr(merged_cells, "ranges", []) if merged_cells is not None else []
    for merged_range in ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row not in visible_row_set or min_col not in visible_column_set:
            continue
        rows = [
            row
            for row in visible_rows
            if min_row <= row <= max_row and row in visible_row_set
        ]
        columns = [
            col
            for col in visible_columns
            if min_col <= col <= max_col and col in visible_column_set
        ]
        if not rows or not columns:
            continue
        starts[(min_row, min_col)] = (len(rows), len(columns))
        for row in rows:
            for col in columns:
                if row == min_row and col == min_col:
                    continue
                covered.add((row, col))
    return starts, covered


def _display_cell_value(cell: Any, cached_sheet: Any | None) -> str:
    """Return the display value for a cell, using cached formula values."""
    raw_value = getattr(cell, "value", None)
    is_formula = isinstance(raw_value, str) and raw_value.startswith("=")
    if is_formula:
        cached_value = (
            cached_sheet.cell(row=cell.row, column=cell.column).value
            if cached_sheet is not None
            else None
        )
        value = cached_value if cached_value is not None else ""
    else:
        value = raw_value
    return _normalize_cell_value(value, getattr(cell, "number_format", "") or "")


def _normalize_cell_value(value: Any, number_format: str = "") -> str:
    """Normalize a worksheet value for stable HTML display."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if "%" in number_format:
            return _format_float(value * 100) + "%"
        return _format_float(value)
    if isinstance(value, int):
        return f"{value:,}" if abs(value) >= 1000 else str(value)
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _format_float(value: float) -> str:
    """Return compact numeric display text."""
    if value.is_integer():
        return f"{int(value):,}" if abs(value) >= 1000 else str(int(value))
    return f"{value:,.6f}".rstrip("0").rstrip(".")


def _html_cell_value(value: str) -> str:
    if not value:
        return "&nbsp;"
    return escape(value).replace("\n", "<br>")


def _column_letter(column_number: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(column_number)


def _row_is_hidden(sheet: Any, row_number: int) -> bool:
    return bool(getattr(sheet.row_dimensions[row_number], "hidden", False))


def _column_is_hidden(sheet: Any, column_number: int) -> bool:
    return bool(
        getattr(sheet.column_dimensions[_column_letter(column_number)], "hidden", False)
    )


def _column_width_px(sheet: Any, column_number: int) -> int:
    dimension = sheet.column_dimensions[_column_letter(column_number)]
    width = getattr(dimension, "width", None)
    if width is None:
        width = getattr(getattr(sheet, "sheet_format", None), "defaultColWidth", 8.43)
    return max(42, min(int(float(width or 8.43) * 7 + 12), 360))


def _row_style(sheet: Any, row_number: int) -> dict[str, str]:
    height = getattr(sheet.row_dimensions[row_number], "height", None)
    if height is None:
        height = getattr(getattr(sheet, "sheet_format", None), "defaultRowHeight", None)
    if not height:
        return {}
    return {"height": f"{max(18, int(float(height) * 1.333))}px"}


def _cell_style(cell: Any) -> dict[str, str]:
    """Return inline CSS approximating core Excel formatting."""
    styles: dict[str, str] = {}
    font = getattr(cell, "font", None)
    if getattr(font, "bold", False):
        styles["font-weight"] = "700"
    if getattr(font, "italic", False):
        styles["font-style"] = "italic"
    if getattr(font, "underline", None):
        styles["text-decoration"] = "underline"
    font_size = getattr(font, "sz", None)
    if font_size:
        styles["font-size"] = f"{float(font_size):.1f}pt"
    font_color = _color_value(getattr(font, "color", None))
    if font_color:
        styles["color"] = font_color

    fill = getattr(cell, "fill", None)
    fill_type = getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)
    if fill_type == "solid":
        fill_color = _color_value(getattr(fill, "fgColor", None))
        if fill_color and fill_color.lower() != "#ffffff":
            styles["background-color"] = fill_color

    alignment = getattr(cell, "alignment", None)
    horizontal = getattr(alignment, "horizontal", None)
    if horizontal:
        styles["text-align"] = str(horizontal)
    vertical = getattr(alignment, "vertical", None)
    if vertical:
        styles["vertical-align"] = str(vertical)
    if getattr(alignment, "wrap_text", False):
        styles["white-space"] = "pre-wrap"

    border = getattr(cell, "border", None)
    for edge in ("left", "right", "top", "bottom"):
        side = getattr(border, edge, None)
        if getattr(side, "style", None):
            styles[f"border-{edge}"] = (
                f"1px solid {_color_value(getattr(side, 'color', None)) or _TABLE_BORDER_COLOR}"
            )
    return styles


def _color_value(color: Any) -> str:
    """Return a CSS hex color for direct RGB openpyxl colors."""
    if color is None:
        return ""
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return ""
    rgb = rgb.strip()
    if len(rgb) == 8:
        alpha, rgb = rgb[:2], rgb[2:]
        if alpha == "00":
            return ""
    if len(rgb) != 6:
        return ""
    return f"#{rgb.upper()}"


def _style_attr(styles: dict[str, str]) -> str:
    if not styles:
        return ""
    value = ";".join(f"{name}:{css_value}" for name, css_value in styles.items())
    return f'style="{escape(value, quote=True)}"'


def _workbook_html_styles() -> str:
    """Return isolated CSS for generated workbook previews."""
    return """
:root {
  color-scheme: light;
  font-family: Arial, Helvetica, sans-serif;
}
* { box-sizing: border-box; }
body {
  background: #f4f6f8;
  color: #111827;
  margin: 0;
}
.workbook-header {
  align-items: center;
  background: #ffffff;
  border-bottom: 1px solid #d8dee6;
  display: flex;
  gap: 20px;
  justify-content: space-between;
  padding: 14px 18px;
  position: sticky;
  top: 0;
  z-index: 4;
}
.workbook-header h1 {
  font-size: 17px;
  line-height: 1.25;
  margin: 0 0 3px;
}
.workbook-header p {
  color: #5f6b7a;
  font-size: 12px;
  margin: 0;
}
.sheet-nav {
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  justify-content: flex-end;
}
.sheet-nav a {
  background: #eef2f6;
  border: 1px solid #d8dee6;
  border-radius: 6px;
  color: #1f4b6e;
  font-size: 12px;
  padding: 6px 8px;
  text-decoration: none;
}
.sheet {
  background: #ffffff;
  border-bottom: 1px solid #d8dee6;
  padding: 18px;
}
.sheet h2 {
  font-size: 15px;
  margin: 0 0 12px;
}
.sheet-scroll {
  overflow: auto;
}
.sheet-grid {
  border-collapse: collapse;
  table-layout: fixed;
}
.sheet-grid th,
.sheet-grid td {
  border: 1px solid #d8dee6;
  font-size: 10pt;
  line-height: 1.25;
  min-height: 18px;
  overflow-wrap: anywhere;
  padding: 3px 5px;
  vertical-align: middle;
}
.sheet-grid th {
  background: #eef2f6;
  color: #475569;
  font-weight: 600;
  position: sticky;
  z-index: 2;
}
.sheet-grid tr:first-child th {
  top: 0;
}
.sheet-grid th:first-child {
  left: 0;
  min-width: 44px;
  text-align: right;
  width: 44px;
}
.sheet-grid .corner {
  left: 0;
  top: 0;
  z-index: 3;
}
.sheet-grid td {
  background: #ffffff;
}
.empty-sheet,
.is-empty p {
  color: #5f6b7a;
  font-size: 13px;
  margin: 0;
}
"""


def _build_xml_transcript_preview(
    *,
    original_bytes: bytes,
    filename: str,
    metadata: dict[str, Any],
) -> SourceDocumentPreview:
    return build_transcript_preview_document(
        original_bytes=original_bytes,
        filename=filename,
        metadata=metadata,
    ).preview


def build_transcript_preview_document(
    *,
    original_bytes: bytes,
    filename: str,
    metadata: Mapping[str, Any] | None = None,
) -> TranscriptPreviewDocument:
    """Build a readable XML transcript PDF and page-aligned text records."""
    effective_metadata = dict(metadata or {})
    if not effective_metadata:
        effective_metadata = _base_metadata(
            filename=filename,
            file_type="xml",
            mime_type="application/xml",
            source_type="",
            file_hash="",
        )

    document = _parse_factset_transcript(original_bytes)
    preview_bytes, pages = _render_transcript_pdf(document, filename)
    if not pages:
        raise ValueError("Transcript XML produced no preview pages.")

    page_metadata = [
        {
            "page_number": page.page_number,
            "preview_page": page.page_number,
            "anchor": f"page-{page.page_number}",
            "title": page.title,
            "section_name": page.section_name,
            "speaker_block_ids": page.speaker_block_ids,
            "speakers": page.speakers,
        }
        for page in pages
    ]
    preview = SourceDocumentPreview(
        preview_mime_type=PDF_MIME_TYPE,
        preview_bytes=preview_bytes,
        preview_metadata={
            **effective_metadata,
            "preview_kind": "styled_transcript_pdf",
            "page_model": "generated_transcript_pdf_pages",
            "transcript_title": document.title,
            "transcript_date": document.transcript_date,
            "companies": document.companies,
            "preview_page_count": len(pages),
            "page_count": len(pages),
            "pages": page_metadata,
        },
    )
    return TranscriptPreviewDocument(
        preview=preview,
        pages=pages,
        title=document.title,
        transcript_date=document.transcript_date,
        companies=document.companies,
    )


def _render_transcript_pdf(
    document: _TranscriptDocument,
    filename: str,
) -> tuple[bytes, list[TranscriptPreviewPage]]:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter, pageCompression=1)
    pdf.setTitle(f"{filename} preview")
    builder = _TranscriptPdfBuilder(
        pdf=pdf,
        page_size=letter,
        filename=filename,
        document=document,
    )

    md_blocks = [
        block for block in document.blocks if block.section_name == _SECTION_MD
    ]
    if md_blocks:
        builder.start_section(_SECTION_MD, "Management Discussion")
        for block in md_blocks:
            builder.draw_speaker_block(block)

    qa_groups = _qa_groups(document.blocks)
    if qa_groups:
        builder.start_section(_SECTION_QA, "Q&A")
        for qa_group_id, qa_blocks in enumerate(qa_groups, start=1):
            builder.draw_exchange_heading(
                f"Q&A Exchange {qa_group_id}: {_qa_title_speaker(qa_blocks)}"
            )
            for block in qa_blocks:
                builder.draw_speaker_block(block)

    emitted_block_ids = {block.speaker_block_id for block in md_blocks} | {
        block.speaker_block_id
        for group in qa_groups
        for block in group
    }
    other_blocks = [
        block
        for block in document.blocks
        if block.speaker_block_id not in emitted_block_ids
    ]
    if other_blocks:
        grouped: dict[str, list[_TranscriptSpeakerBlock]] = {}
        for block in other_blocks:
            grouped.setdefault(block.section_name or "Transcript", []).append(block)
        for section_name, blocks in grouped.items():
            builder.start_section(section_name, section_name)
            for block in blocks:
                builder.draw_speaker_block(block)

    pages = builder.finish()
    pdf.save()
    return buffer.getvalue(), pages


class _TranscriptPdfBuilder:
    """Draw a transcript PDF while collecting page-aligned markdown."""

    def __init__(
        self,
        *,
        pdf: Any,
        page_size: tuple[float, float],
        filename: str,
        document: _TranscriptDocument,
    ) -> None:
        self.pdf = pdf
        self.page_size = page_size
        self.filename = filename
        self.document = document
        self.width, self.height = page_size
        self.page_number = 0
        self.y = 0.0
        self.current_title = ""
        self.current_section = ""
        self.current_lines: list[str] = []
        self.current_speaker_block_ids: set[int] = set()
        self.current_speakers: dict[str, dict[str, str]] = {}
        self.pages: list[TranscriptPreviewPage] = []

    def start_section(self, section_name: str, title: str) -> None:
        """Start a major transcript section on a fresh PDF page."""
        self._new_page(title=title, section_name=section_name, continued=False)
        self._draw_wrapped(
            section_name,
            kind="section",
            markdown_prefix="# ",
            block=None,
        )

    def draw_exchange_heading(self, title: str) -> None:
        """Draw a Q&A exchange heading."""
        self._draw_wrapped(
            title,
            kind="subsection",
            markdown_prefix="## ",
            block=None,
        )

    def draw_speaker_block(self, block: _TranscriptSpeakerBlock) -> None:
        """Draw one speaker block with heading and paragraphs."""
        if block.section_name == _SECTION_QA:
            role = _speaker_role(block)
            speaker_label = f"{role}: {block.participant.display_name}"
        else:
            speaker_label = block.participant.display_name
        self._draw_wrapped(
            speaker_label,
            kind="speaker",
            markdown_prefix="### ",
            block=block,
        )
        for paragraph in block.paragraphs:
            self._draw_wrapped(paragraph, kind="body", markdown_prefix="", block=block)

    def finish(self) -> list[TranscriptPreviewPage]:
        """Finalize the current page and return generated page records."""
        if self.page_number:
            _draw_footer(self.pdf, self.width, self.page_number)
            self._commit_current_page()
            self.pdf.showPage()
        return list(self.pages)

    def _new_page(self, *, title: str, section_name: str, continued: bool) -> None:
        if self.page_number:
            _draw_footer(self.pdf, self.width, self.page_number)
            self._commit_current_page()
            self.pdf.showPage()
        self.page_number += 1
        self.current_title = title
        self.current_section = section_name
        self.current_lines = []
        self.current_speaker_block_ids = set()
        self.current_speakers = {}
        display_title = title + (" (continued)" if continued else "")
        subtitle_parts = [self.filename]
        if self.document.transcript_date:
            subtitle_parts.append(self.document.transcript_date)
        if section_name:
            subtitle_parts.append(section_name)
        subtitle_parts.append(f"preview page {self.page_number}")
        self.pdf.setFillColorRGB(1, 1, 1)
        self.pdf.rect(0, 0, self.width, self.height, fill=1, stroke=0)
        self.y = _draw_page_title(
            self.pdf,
            self.width,
            self.height,
            display_title,
            " | ".join(subtitle_parts),
        )

    def _commit_current_page(self) -> None:
        markdown = "\n".join(line.rstrip() for line in self.current_lines).strip()
        if not markdown:
            markdown = f"# {self.current_title}"
        self.pages.append(
            TranscriptPreviewPage(
                page_number=self.page_number,
                title=self.current_title,
                markdown=markdown,
                section_name=self.current_section,
                speaker_block_ids=sorted(self.current_speaker_block_ids),
                speakers=list(self.current_speakers.values()),
            )
        )

    def _draw_wrapped(
        self,
        text: str,
        *,
        kind: str,
        markdown_prefix: str,
        block: _TranscriptSpeakerBlock | None,
    ) -> None:
        text = _clean_text(text)
        if not text:
            return
        spacing_before, spacing_after, line_height, wrap_width = _pdf_style(kind)
        self._ensure_space(spacing_before + line_height)
        self.y -= spacing_before
        wrapped = textwrap.wrap(text, width=wrap_width) or [text]
        heading_recorded = False
        for line in wrapped:
            self._ensure_space(line_height)
            if block is not None:
                self._note_speaker(block)
            if markdown_prefix and not heading_recorded:
                self.current_lines.append(f"{markdown_prefix}{text}")
                heading_recorded = True
            elif not markdown_prefix:
                self.current_lines.append(line)
            _draw_pdf_text_line(self.pdf, kind, line, _PAGE_MARGIN, self.y)
            self.y -= line_height
        self.y -= spacing_after
        if kind in {"body", "speaker", "subsection", "section"}:
            self.current_lines.append("")

    def _ensure_space(self, required_height: float) -> None:
        if self.page_number and self.y >= _PAGE_MARGIN + required_height:
            return
        if not self.page_number:
            self._new_page(title="Transcript", section_name="", continued=False)
            return
        self._new_page(
            title=self.current_title,
            section_name=self.current_section,
            continued=True,
        )

    def _note_speaker(self, block: _TranscriptSpeakerBlock) -> None:
        self.current_speaker_block_ids.add(block.speaker_block_id)
        key = str(block.speaker_block_id)
        if key not in self.current_speakers:
            self.current_speakers[key] = _transcript_speaker_record(block)


def _pdf_style(kind: str) -> tuple[float, float, float, int]:
    if kind == "section":
        return 0, 10, 15, 80
    if kind == "subsection":
        return 8, 6, 12.5, 86
    if kind == "speaker":
        return 7, 5, 11.5, 92
    return 2, 7, 10.8, 104


def _draw_pdf_text_line(pdf: Any, kind: str, text: str, x: float, y: float) -> None:
    if kind == "section":
        pdf.setFillColorRGB(0.08, 0.12, 0.18)
        pdf.setFont("Helvetica-Bold", 13)
    elif kind == "subsection":
        pdf.setFillColorRGB(0.13, 0.23, 0.34)
        pdf.setFont("Helvetica-Bold", 10.5)
    elif kind == "speaker":
        pdf.setFillColorRGB(0.18, 0.26, 0.36)
        pdf.setFont("Helvetica-Bold", 9.5)
    else:
        pdf.setFillColorRGB(0.1, 0.13, 0.18)
        pdf.setFont("Helvetica", 9)
    pdf.drawString(x, y, _pdf_text(text))


def _transcript_speaker_record(block: _TranscriptSpeakerBlock) -> dict[str, str]:
    return {
        "speaker_block_id": str(block.speaker_block_id),
        "participant_id": block.participant_id,
        "name": block.participant.name,
        "title": block.participant.title,
        "affiliation": block.participant.affiliation,
        "participant_type": block.participant.participant_type,
        "speaker_type_hint": block.speaker_type_hint,
    }


def _draw_page_title(
    pdf: Any,
    width: float,
    height: float,
    title: str,
    subtitle: str,
) -> float:
    pdf.setFillColorRGB(0.1, 0.13, 0.18)
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(_PAGE_MARGIN, height - _PAGE_MARGIN, _pdf_text(_clip(title, 96)))
    pdf.setFillColorRGB(0.35, 0.4, 0.48)
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(
        _PAGE_MARGIN,
        height - _PAGE_MARGIN - 16,
        _pdf_text(_clip(subtitle, 140)),
    )
    pdf.setStrokeColorRGB(0.78, 0.82, 0.87)
    pdf.line(_PAGE_MARGIN, height - _PAGE_MARGIN - 27, width - _PAGE_MARGIN, height - _PAGE_MARGIN - 27)
    return height - _PAGE_MARGIN - 42


def _draw_text_lines(
    pdf: Any,
    lines: list[str],
    *,
    x: float,
    y: float,
    line_height: float,
) -> None:
    pdf.setFillColorRGB(0.1, 0.13, 0.18)
    pdf.setFont("Helvetica", 8.7)
    for line in lines:
        if y < 34:
            break
        pdf.drawString(x, y, _pdf_text(line))
        y -= line_height


def _draw_footer(pdf: Any, width: float, page_number: int) -> None:
    pdf.setFillColorRGB(0.45, 0.5, 0.57)
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(width - _PAGE_MARGIN, 24, f"Preview page {page_number}")


def _parse_factset_transcript(original_bytes: bytes) -> _TranscriptDocument:
    try:
        root = ET.fromstring(original_bytes)
    except ET.ParseError as exc:
        raise ValueError(f"Invalid XML transcript: {exc}") from exc

    namespace = root.tag.split("}")[0] + "}" if root.tag.startswith("{") else ""
    meta = _find(root, namespace, "meta")
    if meta is None:
        raise ValueError("XML transcript meta section is missing.")
    body = _find(root, namespace, "body")
    if body is None:
        raise ValueError("XML transcript body section is missing.")

    title = _element_text(_find(meta, namespace, "title"))
    transcript_date = _element_text(_find(meta, namespace, "date"))
    participants = _parse_participants(meta, namespace)
    companies = _parse_companies(meta, namespace) or _companies_from_participants(
        participants
    )
    blocks = _parse_speaker_blocks(body, namespace, participants)
    if not blocks:
        raise ValueError("No transcript speaker blocks found.")
    return _TranscriptDocument(
        title=title,
        transcript_date=transcript_date,
        participants=participants,
        companies=companies,
        blocks=blocks,
    )


def _parse_companies(meta: ET.Element, namespace: str) -> list[str]:
    companies = []
    companies_el = _find(meta, namespace, "companies")
    if companies_el is not None:
        for company in _findall(companies_el, namespace, "company"):
            company_text = _element_text(company)
            if company_text:
                companies.append(company_text)
    return companies


def _companies_from_participants(
    participants: dict[str, _TranscriptParticipant],
) -> list[str]:
    company_side = [
        participant
        for participant in participants.values()
        if participant.participant_type.lower() in {"executive", "moderator"}
    ]
    companies: list[str] = []
    seen: set[str] = set()
    for participant in company_side or list(participants.values()):
        for value in (participant.affiliation, participant.affiliation_entity):
            company = _clean_text(value)
            key = company.casefold()
            if not company or key in seen:
                continue
            seen.add(key)
            companies.append(company)
    return companies


def _parse_participants(
    meta: ET.Element,
    namespace: str,
) -> dict[str, _TranscriptParticipant]:
    participants: dict[str, _TranscriptParticipant] = {}
    participants_el = _find(meta, namespace, "participants")
    if participants_el is None:
        return participants

    for participant_el in _findall(participants_el, namespace, "participant"):
        participant_id = participant_el.get("id", "").strip()
        if not participant_id:
            continue
        participants[participant_id] = _TranscriptParticipant(
            participant_id=participant_id,
            name=_clean_text(
                participant_el.get("name", "")
                or _element_text(participant_el)
                or "Unknown Speaker"
            ),
            participant_type=participant_el.get("type", "").strip(),
            title=_clean_text(participant_el.get("title", "")),
            affiliation=_clean_text(participant_el.get("affiliation", "")),
            affiliation_entity=_clean_text(
                participant_el.get("affiliation_entity", "")
            ),
        )
    return participants


def _parse_speaker_blocks(
    body: ET.Element,
    namespace: str,
    participants: dict[str, _TranscriptParticipant],
) -> list[_TranscriptSpeakerBlock]:
    blocks: list[_TranscriptSpeakerBlock] = []
    speaker_block_id = 1
    for section_el in _findall(body, namespace, "section"):
        raw_section_name = section_el.get("name", "").strip()
        section_name = _normalize_section_name(raw_section_name)
        for speaker_el in _findall(section_el, namespace, "speaker"):
            participant_id = speaker_el.get("id", "").strip()
            paragraphs = _speaker_paragraphs(speaker_el, namespace)
            if not paragraphs:
                continue
            participant = participants.get(
                participant_id,
                _TranscriptParticipant(
                    participant_id=participant_id,
                    name="Unknown Speaker",
                ),
            )
            blocks.append(
                _TranscriptSpeakerBlock(
                    speaker_block_id=speaker_block_id,
                    section_name=section_name,
                    raw_section_name=raw_section_name,
                    participant_id=participant_id,
                    speaker_type_hint=speaker_el.get("type", "").strip().lower(),
                    participant=participant,
                    paragraphs=paragraphs,
                )
            )
            speaker_block_id += 1
    return blocks


def _speaker_paragraphs(speaker_el: ET.Element, namespace: str) -> list[str]:
    paragraphs = []
    plist = _find(speaker_el, namespace, "plist")
    paragraph_parent = plist if plist is not None else speaker_el
    for paragraph_el in _findall(paragraph_parent, namespace, "p"):
        paragraph = _element_text(paragraph_el)
        if paragraph:
            paragraphs.append(paragraph)
    return paragraphs


def _qa_groups(
    blocks: list[_TranscriptSpeakerBlock],
) -> list[list[_TranscriptSpeakerBlock]]:
    qa_blocks = [block for block in blocks if block.section_name == _SECTION_QA]
    groups: list[list[_TranscriptSpeakerBlock]] = []
    current_group: list[_TranscriptSpeakerBlock] = []
    for block in qa_blocks:
        starts_question = block.speaker_type_hint == "q"
        if starts_question and current_group:
            groups.append(current_group)
            current_group = [block]
            continue
        current_group.append(block)
    if current_group:
        groups.append(current_group)
    return groups


def _speaker_role(block: _TranscriptSpeakerBlock) -> str:
    if block.speaker_type_hint == "q":
        return "Question"
    if block.speaker_type_hint == "a":
        return "Answer"
    participant_type = block.participant.participant_type.casefold()
    if participant_type == "analyst":
        return "Question"
    if participant_type in {"executive", "moderator"}:
        return "Answer"
    return "Speaker"


def _qa_title_speaker(blocks: list[_TranscriptSpeakerBlock]) -> str:
    question_block = next(
        (
            block
            for block in blocks
            if block.speaker_type_hint == "q"
            or block.participant.participant_type.casefold() == "analyst"
        ),
        None,
    )
    return (question_block or blocks[0]).speaker_name if blocks else "Unknown Speaker"


def _normalize_section_name(section_name: str) -> str:
    normalized = section_name.casefold()
    if "management" in normalized and "discussion" in normalized:
        return _SECTION_MD
    if "question" in normalized or "q&a" in normalized:
        return _SECTION_QA
    return section_name.strip() or "Transcript"


def _find(parent: ET.Element, namespace: str, tag: str) -> ET.Element | None:
    return parent.find(f"{namespace}{tag}")


def _findall(parent: ET.Element, namespace: str, tag: str) -> list[ET.Element]:
    return list(parent.findall(f"{namespace}{tag}"))


def _element_text(element: ET.Element | None) -> str:
    if element is None:
        return ""
    return _clean_text(" ".join(element.itertext()))


def _clean_text(text: str) -> str:
    return _WHITESPACE_RE.sub(" ", text or "").strip()


def _clip(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[: max(0, max_chars - 3)] + "..."


def _pdf_text(text: str) -> str:
    return str(text).encode("latin-1", errors="replace").decode("latin-1")
